import openpyxl
import re

# load excel with its path
wrkbk = openpyxl.load_workbook("HCA35_Index_clean.xlsx")
  
sh = wrkbk.active
#s1 = wb.create_sheet("Mysheet")
dateArr = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
abbrDates = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
  
# iterate through excel and display data
for i in range(1, sh.max_row+1):
      
    dateOfAdju = sh.cell(row=i, column=9).value
    dateOfCapture = sh.cell(row=i, column=10).value
    dateOfAdjuSplit = []
    dateOfCaptureSplit = []
    if (type(dateOfAdju) == str):
        dateStr = dateOfAdju.strip()
        #print(dateStr)
        dateOfAdjuSplit = re.split(" |, |,| , ", dateStr)
    if (len(dateOfAdjuSplit) == 3):
        if(dateOfAdjuSplit[0] in dateArr and dateOfAdjuSplit[1].isdigit()):
            newDate = dateOfAdjuSplit[1] + " " + dateOfAdjuSplit[0] + " " + dateOfAdjuSplit[2]
            sh.cell(row=i, column=9, value=newDate)
            #print(newDate)
    if (type(dateOfCapture) == str):
        dateStr = dateOfCapture.strip()
        #print(dateStr)
        dateOfCaptureSplit = re.split(" |, |,| , ", dateStr)
    if (len(dateOfCaptureSplit) == 3):
        if(dateOfCaptureSplit[0] in dateArr and dateOfCaptureSplit[1].isdigit()):
            newDate = dateOfCaptureSplit[1] + " " + dateOfCaptureSplit[0] + " " + dateOfCaptureSplit[2]
            sh.cell(row=i, column=10, value=newDate)
wrkbk.save('HCA.xlsx')
