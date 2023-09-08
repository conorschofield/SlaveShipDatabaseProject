import mysql.connector
from openpyxl import load_workbook, Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from collections import Counter
import re
import datetime
from hcadateparser.dateparser import validateResult, parseDate
from hcaconcordance.concordance import generateConcordance
from datetime import datetime
from sys import argv, exit

if len(argv) < 2:
    print("Usage: betterimport.py <case image directory>")
    exit(-1)

caseImageDirectory = argv[1]

print ("# HCA 35 Import Report")

INPUT_WORKBOOK_FILENAME = 'workbooks/HCA 35 Index Merged as Separate Sheets.xlsx'

EXCEL_OUTPUT_FILE = None

# Whether we should just write the cleaned data to an Excel
# Uncomment the next line and comment the line afterwards to enable
# EXCEL_OUTPUT_FILE = 'HCA 35 Index (Cleaned and Merged) DateFix.xlsx'

# Assume that all rows in volumes that follow 77 (but not including 77) are all
# Indian Ocean unless the Court value indicates otherwise. Beyond this point,
# highlighting no longer provides meaningful data.
LAST_VOLUME_WITH_NOTABLE_ATLANTIC_OCEAN_CASES = 77

# Assume that these volumes don't have highlighting for ocean data, and instead
# only use the "Ocean" column.
NO_OCEAN_HIGHLIGHTING_VOLUMES = {43, 45, 71, 72, 73, 74, 79, 80}

# Purple & pink highlighting is ignored.
IGNORED_HIGHLIGHT_COLORS = {"FFE28CF5", "FFF5BAEB"}

# Whether warnings should be generated if there exists highlighting
# exclusively on the Case or Red Number columns. By default this
# highlighting is ignored.
WARN_INSIGNIFICANT_HIGHLIGHTING = False

# Someone put their timesheet info into the spreadsheet and forgot to
# remove it. We'll discard it when importing.
IGNORED_UNNAMED_COLUMN_DATA = {"2 Hours", "40/121"}

ATLANTIC_OCEAN_COURTS = {"Sierra Leone", "Saint Helena", "Rio de Janeiro", "Jamaica", \
    "Havana", \
    "Tortola"}
INDIAN_OCEAN_COURTS = {"Mauritius", "Zanzibar", "Cape of Good Hope"}

# Try to match either:
# - a single number
# - a bunch of numbers separated with commas or slashes
# - a bunch of number separated with spaces
# - a pair of numbers separated with an ampersand
#
# If you don't understand this, try pasting it in to regex101.com.
rednumber_pair_regex = "([0-9]+\s*\&\s*[0-9]+)"
number_regex = "([0-9]+)(\s*(a|\(a\)))?"
rednumber_pattern = re.compile("^(" + rednumber_pair_regex + "|((" + number_regex + "(\s*(\/|\,\s*)\s*|\s+))*" + number_regex + "(\/|\,\s*)?))$")
court_pattern = re.compile("^[a-zA-Z\s.]+$")

IMAGE_NAME_PATTERN = re.compile('^DSC[0-9]{5}$')

def normalize_ocean(ocean):
    if ocean is None:
        return None

    if ocean == "Atlantic" or ocean == "Indian":
        return ocean
    elif ocean == "Indian Ocean":
        return "Indian"
    elif ocean == "South China Sea":
        # Note: this is an edge case, row 2789
        return "South China Sea"
    else:
        raise ValueError("Unrecognized ocean \"%s\"" % ocean)

warn_reasons = Counter()
courts = Counter()
commissions = Counter()
warnings_kept = set()
discarded = set()
discard_reasons = Counter()
num_imported = 0
num_empty = 0

def get_row_values(row):
    # Get just the cell values without formatting, or None if the cell is empty
    row_values = [cell.value for cell in row]

    # Remove unnecessary whitespace
    row_values = [v.strip() if isinstance(v, str) else v for v in row_values]

    # Replace cells that are only whitespace with None values
    row_values = [v if not isinstance(v, str) or len(v) != 0 else None for v in row_values]

    # Figure out what cells have non-None values
    nonempty_cols = [(idx, v) for (idx, v) in enumerate(row_values) if v is not None]    

    return (row_values, nonempty_cols)

def capture_base_schema(row):
    base_schema = dict()
    base_unnamed_cols = []

    (row_values, nonempty_cols) = get_row_values(row)

    for (idx, key) in enumerate(row_values):
        if key is not None:
            base_schema[key] = idx
        else:
            base_unnamed_cols.append(idx)

    #print("Captured base schema %s" % base_schema)

    return (base_schema, base_unnamed_cols)

def get_volume_schema(volume, base_schema, base_unnamed_cols):
    schema = base_schema
    unnamed_cols = base_unnamed_cols

    if volume == 62:
        # Volume 62 inserts an extra column at index 13 for the "Captain", which
        # we currently ignore.
        schema = dict()
        schema["Captain"] = 13
        unnamed_cols = [u for u in base_unnamed_cols if u != 22]

        for key in base_schema:
            col_idx = base_schema[key]
            # The column was inserted at index 13, so the base schema needs to be shifted
            # by one column.
            if col_idx >= 13:
                col_idx += 1
            schema[key] = col_idx
    
    if volume == 87:
        img_cols = [22, 23, 24]
        unnamed_cols = [u for u in base_unnamed_cols if u not in img_cols]
        for idx, col in enumerate(img_cols):
            schema['ImageName' + str(idx)] = col

    if volume in [84, 85]:
        unnamed_cols = [u for u in base_unnamed_cols if u != 22]
        schema['Transcriber Note'] = 22

    return (schema, unnamed_cols)

# Some rows have a single, non-empty column that does not contain significant data and is just
# copy-pasted data. These are semantically empty.
def is_row_empty(nonempty_cols):
    if not nonempty_cols:
        return True

    if (len(nonempty_cols) == 1 and \
            (nonempty_cols[0][0] == base_schema["Ocean"] or nonempty_cols[0][0] == base_schema["Record Group"])):
        return True

    if (len(nonempty_cols) == 2 and \
            (nonempty_cols[0][0] == base_schema["Record Group"] or nonempty_cols[1][0] == base_schema["Volume"])):
        return True

    return False

class InconsistentHighlightingError(Exception): pass

def infer_ocean_from_highlighting(row):
    num_yellow = 0
    num_clear = 0
    num_nonempty = 0
    highlighting_is_insigificant = True

    for (col_idx, cell) in enumerate(row):
        # First two columns might not be highlighted. This seems to be an artifact of
        # how the data was copy-pasted into the spreadsheet?
        if cell.value is None or col_idx <= 1:
            continue

        insignificant = col_idx == base_schema["Case (ship)"] or col_idx == base_schema["Red number"]
        num_nonempty += 1
        
        color = cell.fill.start_color.value
        
        if color == "FFFFFF00":
            if not insignificant:
                highlighting_is_insigificant = False
            num_yellow += 1
        elif color == "00000000":
            num_clear += 1
        elif volume > LAST_VOLUME_WITH_NOTABLE_ATLANTIC_OCEAN_CASES:
            # Ignore all highlighting after this volume since it has no significant meaning.
            pass
        elif color in IGNORED_HIGHLIGHT_COLORS:
            # Ignore this highlighting.
            pass
        elif not insignificant:
            print("      [Keeping row] [row " + str(row_number) + "] Unrecognized highlight color %s (treating as indeterminate color)" % (color))
            warnings_kept.add(row_number)
            warn_reasons["unrecognized highlight color"] += 1
        elif insignificant and WARN_INSIGNIFICANT_HIGHLIGHTING:
            print("      [Keeping row] [row " + str(row_number) + "] Ignoring insignificant highlighting of unrecognized color")
            warnings_kept.add(row_number)
            warn_reasons["insignificant unrecognized highlighting"] += 1

    inferred_ocean = None

    # note: Ocean column is sometimes missing highlighting
    if volume > LAST_VOLUME_WITH_NOTABLE_ATLANTIC_OCEAN_CASES:
        # TODO: Verify Indian Ocean
        inferred_ocean = None
    elif num_yellow > 0 and num_yellow >= num_nonempty - 1 and not highlighting_is_insigificant:
        inferred_ocean = "Indian"
    elif num_yellow == 0 or highlighting_is_insigificant:
        inferred_ocean = "Atlantic"
    else:
        raise InconsistentHighlightingError

    return inferred_ocean

class MalformedRedNumbersError(Exception):
    def __init__(self, court, red_numbers):
        self.court = court
        self.red_numbers = red_numbers

class CantParseCourtAndRedNumbersError(Exception):
    def __init__(self, maybe_court, maybe_red_numbers):
        self.maybe_court = maybe_court
        self.maybe_red_numbers = maybe_red_numbers

class RedNumberDataLossError(): pass

def handle_court_and_red_numbers(maybe_red_numbers, maybe_court):
    # Hardcoded case of a red number range - general handling would be complex and potentially error-prone.
    if maybe_red_numbers == "3342-3345":
        maybe_red_numbers = "3342, 3343, 3344, 3345"
    
    if maybe_red_numbers == "7038. 7033":
        maybe_red_numbers = "7038, 7033"

    # Various ways of indicating that the data is absent
    if maybe_red_numbers == "(No Red No.)" \
        or maybe_red_numbers == "N/A" \
        or maybe_red_numbers == "page missing" \
        or maybe_red_numbers == "NO RED #":
        maybe_red_numbers = None

    if maybe_court == "_":
        maybe_court = None

    # Error in the data that's trivial to fix.
    if maybe_court == "yes, British and Spanish" \
          and row_values[schema["Mixed Commission?"]] is None:
        maybe_court = None
        row_values[schema["Mixed Commission?"]] = "yes, British and Spanish"

    red_numbers = None
    court = None

    if maybe_red_numbers is not None and \
        (isinstance(maybe_red_numbers, int) \
            or isinstance(maybe_red_numbers, float) \
            or rednumber_pattern.match(str(maybe_red_numbers)) is not None):
        red_numbers = maybe_red_numbers
        court = maybe_court
    elif maybe_court is not None and \
        (isinstance(maybe_court, int) \
            or isinstance(maybe_court, float) \
            or rednumber_pattern.match(str(maybe_court)) is not None):
        red_numbers = maybe_court
        court = maybe_red_numbers
    elif maybe_court is None and maybe_red_numbers is None:
        court = None
        red_numbers = None
    elif maybe_court is not None and court_pattern.match(str(maybe_court)) is not None:
        court = maybe_court
        red_numbers = maybe_red_numbers

        if red_numbers is not None:
            raise MalformedRedNumbersError(court, red_numbers)
    elif maybe_red_numbers is not None and court_pattern.match(str(maybe_red_numbers)) is not None:
        court = maybe_red_numbers
        red_numbers = maybe_court

        if red_numbers is not None:
            raise MalformedRedNumbersError(court, red_numbers)
    else:
        raise CantParseCourtAndRedNumbersError(maybe_court, maybe_red_numbers)

    #if court in COURT_SPELLING_FIXES:
    #    court = COURT_SPELLING_FIXES[court]

    # Red numbers are sometimes just strung together without spaces, resulting in MASSIVE numbers when parsed.
    # We need to check if the number is above ~13k, and if so... warn so we can special case it /shrug
    if (isinstance(red_numbers, float) or isinstance(red_numbers, int)) and red_numbers >= 100000:
        red_numbers = RedNumberDataLossError()

    return (court, red_numbers)

def massageStat(count):
    # formatting: "85 (healthy), 30 (sick) "
    if count is None or type(count) is int:
        return count

    count = count.strip()

    if count == "not stated" or count == "_":
        return None
    
    if count == "[illegible]" or count == "illegible":
        return None

    if len(count) == 0:
        return 0

    if "(healthy)" in count and count.endswith(" (sick)"):
        parts = count.split(" (healthy), ")
        healthy = parts[0]
        sick = parts[1]
        sick = sick[0:len(sick) - len(" (sick)")]
        return int(healthy) + int(sick)

    return count

totalNonIntegers = set()

def massageStats(totalS, menS, womenS, boysS, girlsS):
    if type(menS) is str:
        if menS.endswith(" males"):
            menS = menS[0:len(menS) - len(" males")]
        elif menS.endswith("(males)"):
            menS = menS[0:len(menS) - len("(males)")]
        elif menS.endswith("(total males)"):
            menS = menS[0:len(menS) - len("(total males)")]
        elif menS.endswith("(M)"):
            menS = menS[0:len(menS) - len("(M)")]
        elif "Between 500 and 600 African captives" == menS:
            menS = 500

    if type(womenS) is str:
        if womenS.endswith(" females"):
            womenS = womenS[0:len(womenS) - len(" females")]
        elif womenS.endswith("(females)"):
            womenS = womenS[0:len(womenS) - len("(females)")]
        elif womenS.endswith("(total females)"):
            womenS = womenS[0:len(womenS) - len("(total females)")]
        elif womenS.endswith("(F)"):
            womenS = womenS[0:len(womenS) - len("(F)")]

    if type(boysS) is str:
        # data weirdness
        if boysS.endswith(" children"):
            boysS = boysS[0:len(boysS) - len(" children")]

        if boysS.endswith(" children "):
            boysS = boysS[0:len(boysS) - len(" children ")]

        if boysS.endswith(" child"):
            boysS = boysS[0:len(boysS) - len(" child")]

        # for some reason, sometimes the boys number has brackets around it
        # not sure why.
        if boysS.startswith("[") and boysS.endswith("]"):
            boysS = boysS[1:len(boysS)-1]

    totalS = massageStat(totalS)
    menS = massageStat(menS)
    womenS = massageStat(womenS)
    boysS = massageStat(boysS)
    girlsS = massageStat(girlsS)

    if menS == "13+":
        menS = "13"

    if womenS == "10+":
        womenS = "10"

    if boysS == "14+":
        boysS = "14"
    
    if girlsS == "3+":
        girlsS = "3"

    if girlsS == "14 (In addition to 6 infants)":
        girlsS = "20"

    if girlsS == "14 (in addition to 6 infants)":
        girlsS = "20"

    if boysS == '166(164 Healthy and 2 Sickly)':
        boysS = "166"

    if menS == "53(50 Healthy, 3 Sickly, and 7 dead":
        menS = "53"
    
    if womenS == "19 (1 dead)":
        womenS = "19"

    if boysS == "37(35 Healthy, 2 Sickly, 3 dead)":
        boysS = "37"

    if girlsS == "11(10 Healthy, 1 Sickly, 0 dead)":
        girlsS = "11"
    
    if menS == "159(160)":
        menS = "159"
    
    if menS == "34(41)":
        menS = "34"

    if womenS == "37(38)":
        womenS = "37"

    if boysS == "69(86)":
        boysS = "69"

    if girlsS == "29(38)":
        girlsS = "29"

    if menS == "112(129)":
        menS = "112"

    if menS == "235(4 sickly)":
        menS = "235"

    if boysS == "152(6 sickly)":
        boysS = "152"

    try:
        total = int(totalS) if totalS is not None else None
    except ValueError:
        totalNonIntegers.add(totalS)
        total = None

    men = int(menS) if menS is not None else None
    women = int(womenS) if womenS is not None else None
    boys = int(boysS) if boysS is not None else None
    girls = int(girlsS) if girlsS is not None else None

    return (total, men, women, boys, girls)

class SqlImportBatch:
    captors = []
    seen_captors = dict()
    locations = []
    seen_locations = dict()
    sql_captors = []
    sql_locations = []
    sql_dates_cap = []
    sql_dates_cap_verbatim = []
    sql_dates_adj = []
    sql_dates_adj_verbatim = []
    cases = []
    stats = []
    case_counter = 0

    def insertParsedDates(self, sql_dates, sql_dates_verbatim, uniq, parsed):
        for date in parsed['dates']:
            day = date['day']
            month = date['month']
            year = date['year']

            if day is None:
                day = 1
            
            if month is None:
                month = 1

            if year is None:
                year = 1

            if month == 2 and day >= 29:
                day = 1
                
            if month is not None and month == 4 or month == 6 or month == 9 or month == 11:
                if day is not None and day > 30:
                    problems.append(f'validation failure: day {day} in date "{original}" is out of range for given month "{month}". Please re-check against the HCA 35 originals.')
                    day = None
            if month is not None and month == 2:
                if day is not None and day > 28:
                    problems.append(f'validation failure: day {day} in date "{original}" is out of range for given month "{month}". Please re-check against the HCA 35 originals.')
                    day = None

            sql_dates.append((uniq, '%04d-%02d-%02d' % (year, month, day)))

        if parsed['verbatim'] is not None:
            sql_dates_verbatim.append((uniq, parsed['verbatim']))

    def insert(self, volume, start_page, end_page, case_name, red_numbers, court, mixed, ocean, adjudicationDateParsed, \
               captureDateParsed, row_captor, row_location, numSlaves, men, women, boys, girls, register, notes, transcriber_notes):
        uniq = self.case_counter
        self.case_counter += 1
        
        # check if Captor has been seen
        if row_captor is not None and (row_captor not in self.seen_captors):
            ident = len(self.captors) + 1
            self.seen_captors[row_captor] = ident
            self.captors.append(row_captor)

            self.sql_captors.append((ident, row_captor))

        # check if Location has been seen
        if row_location is not None and (row_location not in self.seen_locations):
            ident = len(self.locations) + 1
            self.seen_locations[row_location] = ident
            self.locations.append(row_location)

            self.sql_locations.append((ident, row_location))

        self.cases.append([uniq, volume, start_page, end_page, case_name, red_numbers, court, mixed, ocean,
                           self.seen_captors[row_captor] if row_captor is not None else None,
                           self.seen_locations[row_location] if row_location is not None else None,
                           register, notes, transcriber_notes])
        
        if adjudicationDateParsed[1] is not None:
            self.insertParsedDates(self.sql_dates_adj, self.sql_dates_adj_verbatim, uniq, adjudicationDateParsed[1])

        if captureDateParsed[1] is not None:
            self.insertParsedDates(self.sql_dates_cap, self.sql_dates_cap_verbatim, uniq, captureDateParsed[1])

        self.stats.append([uniq, men, women, boys, girls, 0, numSlaves])
    
    # Caller TODO: Error handling
    def connect_and_execute(self, user, password, host, port, database):
        #print(self.sql_dates_adj)
        #print(self.sql_dates_cap)
        #print(self.sql_dates_adj_verbatim)
        #print(self.sql_dates_cap_verbatim)
        #print(len(self.sql_dates_adj))
        #print(len(self.sql_dates_cap))
        #print(len(self.sql_dates_adj_verbatim))
        #print(len(self.sql_dates_cap_verbatim))

        cnx = mysql.connector.connect(user=user,
                                      password=password,
                                      host=host,
                                      port=port)

        cursor = cnx.cursor(prepared=False)

        sql_commands = None
        with open('create.sql', 'r') as file:
            sql_commands = file.read().split(';')[:-1]
        
        for sql_command in sql_commands:
            print(sql_command)
            cursor.execute(sql_command)

        cursor.close()
        cursor = cnx.cursor(prepared=True)

        try:
            statement = "INSERT INTO Captor VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_captors)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO Location VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_locations)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO Cases (Uniq, Volume, StartPage, EndPage, CaseN, RedNumber, Court, Mixed, "\
                    "Ocean, Captor, Location, Register, Notes, TranscriberNotes) VALUES " \
                "(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
            cursor.executemany(statement, self.cases)
        except mysql.connector.Error as err:
            raise err
        
        try:
            statement = "INSERT INTO SlaveStats (Uniq, Men, Women, Children_m, Children_f, Died, Total) VALUES "\
                " (%s, %s, %s, %s, %s, %s, %s);"
            cursor.executemany(statement, self.stats)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO CaptureDates (CaseId, CaptureDate) VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_dates_cap)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO AdjudicationDates (CaseId, AdjudicationDate) VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_dates_adj)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO CaptureDatesVerbatim (CaseId, CaptureDateVerbatim) VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_dates_cap_verbatim)
        except mysql.connector.Error as err:
            raise err

        try:
            statement = "INSERT INTO AdjudicationDatesVerbatim (CaseId, AdjudicationDateVerbatim) VALUES (%s, %s);"
            cursor.executemany(statement, self.sql_dates_adj_verbatim)
        except mysql.connector.Error as err:
            raise err

        try:
            concordances = generateConcordance(caseImageDirectory)
            statement = "INSERT INTO ImageConcordance (Volume, PageNumber, ImagePath, Weight, OtherPageNumber) VALUES (%s, %s, %s, %s, %s);"
            cursor.executemany(statement, concordances)
        except mysql.connector.Error as err:
            raise err

        cnx.commit()

def highlightedCell(ws, value, highlight):
    if highlight is not None:
        cell = WriteOnlyCell(ws, value)
        cell.fill = PatternFill("solid", Color(highlight), Color(highlight))
        return cell
    else:
        return value

class ExcelImportBatch:
    def __init__(self, filename):
        self.out_wb = Workbook(write_only=True)
        self.out_ws = self.out_wb.create_sheet()
        self.out_ws.freeze_panes = "A2"
        self.out_ws.print_title_rows = '1:1'
        self.out_filename = filename
        self.volumes = dict()

        titles = ["Record Group", "Vol.", "Start Page", "End Page", "Case (ship)", "Red number", \
            "Court", "Mixed Commission?", "Ocean ", "Date of Adjudication", "Date of Capture", "Capturing ship/captor", \
            "Place of Capture", "Number of Slaves", "Men", "Women", "Boys", "Girls", "Register?", "Notes", "Transcriber Notes"]
        title_cells = [WriteOnlyCell(self.out_ws, val) for val in titles]

        for title_cell in title_cells:
            title_cell.font = Font(bold=True)

        self.out_ws.append(title_cells)


    def insert(self, volume, start_page, end_page, case_name, red_numbers, court, mixed, ocean, adjudicationDateParsed, \
               captureDateParsed, row_captor, row_location, numSlaves, men, women, boys, girls, register, notes, transcriber_notes):
        adjudicationDate = adjudicationDateParsed[0]
        captureDate = captureDateParsed[0]
        basevals = ["HCA 35", volume, start_page, end_page, case_name, red_numbers, court, mixed, ocean, adjudicationDate, \
            captureDate, row_captor, row_location, numSlaves, men, women, boys, girls, register, notes, transcriber_notes]
        highlightColor = None if ocean == "Atlantic" else "FFFFFF00"
        cells = [highlightedCell(self.out_ws, val, highlightColor) for val in basevals]
        self.volumes.setdefault(volume, []).append(cells)

    def connect_and_execute(self, user, password, host, port, database):
        for volume in sorted(self.volumes):
            for row in self.volumes[volume]:
                self.out_ws.append(row)

        self.out_wb.save(self.out_filename)

num_sheets_imported = 0

if EXCEL_OUTPUT_FILE is not None:
    importBatch = ExcelImportBatch(EXCEL_OUTPUT_FILE)
else:
    importBatch = SqlImportBatch()

workbook_filename = INPUT_WORKBOOK_FILENAME
wb = load_workbook(
        filename = workbook_filename, 
        read_only=True)

for sheet in wb.worksheets:
    base_schema=None
    base_unnamed_cols=None

    try:
        (base_schema, base_unnamed_cols) = capture_base_schema(next(sheet.rows))

        if "Vol." not in base_schema:
            base_schema["Vol."] = base_schema["Volume"]
    except StopIteration:
        print("Warning: " + sheet.title + " appears to be empty, it is missing a header line.")
    
    num_sheets_imported += 1


    prev_volume = None
    prev_end_page = None

    for (row_number, row) in enumerate(sheet.rows, 1):
        # note: row is a tuple of cells
    
        # First row was the header row, skip it.
        if row_number == 1:
            continue
    
        # Get the row values, and the list of nonempty column indices
        (row_values, nonempty_cols) = get_row_values(row)
    
        # Check if the row is semantically empty before performing further processing.
        # We often find groups of empty rows between volumes, an artifact of how new
        # volumes were copy-pasted in to the data set into a pre-made gap.
        if is_row_empty(nonempty_cols):
            num_empty += 1
            continue
        
        # Get the volume and corresponding schema straight away.
        #
        # This SHOULD be the same Excel column on all rows. This is indeed the case on
        # the current data set I'm working with.
        #
        # Different people performed data entry for different groups of volumes,
        # so different volumes have different quirks. We often need to differentiate
        # based on the volume to work around these quirks.
        volume = int(row_values[base_schema["Vol."]])
        (schema, unnamed_cols) = get_volume_schema(volume, base_schema, base_unnamed_cols)
    
        def report_column_problem(reason_id, columns, reason_explanation):
            where = f'row #{str(row_number)} of volume #{volume} in sheet "{sheet.title}" of workbook "{workbook_filename}"'

            columns_str = ', '.join(columns)

            print(f'- Transcription problem: {reason_id}')
            print(f'    - Location: {where}')
            print(f'    - Columns needing correction: {columns_str}')

            if (isinstance(reason_explanation, list)):
                print('    - Explanation:')
                for line in reason_explanation:
                    print(f'        - {line}')
            else:
                print(f'    - Explanation: {reason_explanation}')
            warnings_kept.add(row_number)
            warn_reasons[reason_id] += 1

        # Find data that exists in unnamed columns, and warn about it.
        # This might indicate unwritten & implicit parts of the schema
        for unnamed_col in unnamed_cols:
            unnamed_val = row_values[unnamed_col]
            if unnamed_val is not None and unnamed_val not in IGNORED_UNNAMED_COLUMN_DATA:
                column_name = chr(ord('A') + unnamed_col) if unnamed_col < 26 else "(#" + str(unnamed_col + 1) + ")"
                print("      [Keeping row] [row " + str(row_number) + " in volume " + str(volume) + "] Unnamed column "\
                    "%s had data \"%s\"" % (column_name, unnamed_val))
                warnings_kept.add(row_number)
                warn_reasons["unnamed data"] += 1
    
        # Ensure that the record group matches HCA 35 like we'd expect. This is largely a redundant
        # column, but if we're discarding it we want to make sure it doesn't take on any unexpected values.
        record_group = row_values[schema["Record Group"]]
    
        if record_group != "HCA 35" and record_group != "HCA35":
            print("   [Discarding row] [row " + str(row_number) + " in volume " + str(volume) + "] Record with group "\
                "\"" + str(record_group) + "\" is not from HCA 35.")
            discarded.add(row_number)
            discard_reasons["wrong record group"] += 1
            continue
    
        # Try to infer the ocean from highlighting.
        try:
            inferred_ocean = infer_ocean_from_highlighting(row)
        except InconsistentHighlightingError:
            print("   [Discarding row] [row " + str(row_number) + "] Could not infer if row was Indian Ocean or "\
                "Atlantic Ocean due to inconsistent highlighting in row")
            discarded.add(row_number)
            discard_reasons["inconsistent highlighting"] += 1
            continue
    
        start_page = row_values[schema["Start Page"]]
        start_page = int(start_page) if start_page is not None else None
        end_page = row_values[schema["End Page"]]
        end_page = int(end_page) if end_page is not None else None
        case_name = str(row_values[schema["Case (ship)"]])
    
        if start_page is None and end_page is None:
            print("   [Discarding row] [row %d] Missing page range! "\
                "You must specify the start page, end page, or both." % row_number)
            discarded.add(row_number)
            discard_reasons["missing page range"] += 1
            continue
        elif start_page is None:
            reason = f"The start page was not written, so it is assumed to be same as the end page ({end_page})."\
                " It is best to state this explicitly instead to avoid ambiguity."
            report_column_problem("assuming start page", ["Start Page"], reason)
            start_page = end_page
        elif end_page is None:
            reason = f"The end page was not written, so it is assumed to be same as the start page ({start_page})."\
                " It is best to state this explicitly instead to avoid ambiguity."
            report_column_problem("assuming end page", ["End Page"], reason)
            end_page = start_page
        elif end_page < start_page:
            reason = f"End page {end_page} is before start page {start_page}. This is probably a typo."
            report_column_problem("end page before start page", ["Start Page", "End Page"], reason)
            end_page = None
            start_page = None
    
        if prev_volume is None or prev_volume != volume:
            prev_volume = volume
            prev_end_page = None
        
        if end_page is not None:
            if prev_end_page is None:
                prev_end_page = end_page
            elif end_page < prev_end_page:
                print(f"alert: Page range went backwards in volume {volume}! This is robably invalid.")

        # Parse, validate, and clean the court & red number fields.
        # These are sometimes swapped, sometimes have misspellings, and sometimes have
        # wacko formatting, so a lot of logic is in this function.
        try:
            (court, red_numbers) = handle_court_and_red_numbers(\
                row_values[schema["Red number"]], row_values[schema["Court"]])
        except MalformedRedNumbersError as e:
            reason = "Malformed red number set \"%s\" (court is \"%s\")" % (e.red_numbers, e.court)
            report_column_problem("malformed red number(s)", ["Red Number"], reason) 
            red_numbers = None
            continue
        except CantParseCourtAndRedNumbersError as e:
            reason = f"Can't tell if Court and Red Number are good -"\
                    f" given \"{e.maybe_red_numbers}\" and \"{e.maybe_court}\""
            report_column_problem("unable to parse court & red number", ["Court", "Red Number"], reason)
            red_numbers = None
            court = None

        if isinstance(red_numbers, RedNumberDataLossError):
            reason = "Red number(s) is/are way too big! This usually means that multiple " \
                    "red numbers were entered without space separation, or Excel thought that it was a single big " \
                    "number. As a result, it has been unpredicably rounded and truncated, causing unrecoverable " \
                    "data loss. Use slashes to separate multiple red numbers instead of commas or similar."
            report_column_problem("red number data loss", ["Red Number"], reason) 
            red_numbers = None
    
        courts[court] += 1
        commissions[row_values[schema["Mixed Commission?"]]] += 1
        raw_ocean = row_values[schema["Ocean"]]
        normalized_ocean = normalize_ocean(raw_ocean)
    
        if volume in NO_OCEAN_HIGHLIGHTING_VOLUMES:
            inferred_ocean = normalized_ocean
        elif volume > LAST_VOLUME_WITH_NOTABLE_ATLANTIC_OCEAN_CASES:
            inferred_ocean = normalized_ocean
        elif normalized_ocean is not None and normalized_ocean != inferred_ocean: 
            reason = "Inconsistent ocean data: Ocean column value \"" + raw_ocean + \
                "\" (normalized form: \"" + normalized_ocean + "\") does not match ocean value \"" + \
                    inferred_ocean + "\" inferred from highlighting"
            report_column_problem("inconsistent ocean data", ["Ocean", "(any highlighted columns)"], reason)
            inferred_ocean = None
    
        ocean = inferred_ocean
    
        (total, men, women, boys, girls) = massageStats(
            row_values[schema["Number of Slaves"]],
            row_values[schema["Men"]],
            row_values[schema["Women"]],
            row_values[schema["Boys"]],
            row_values[schema["Girls"]])
    
        # Consider checking...
        #total = men if men is not None else 0 + \
        #        women if women is not None else 0 + \
        #        boys if boys is not None else 0 + \
        #        girls if girls is not None else 0
        
        #if men is None and women is None and boys is None and girls is None:
        #    total = None

        def handleDate(date_name, date_column_name):
            verbatim = row_values[schema[date_column_name]]
            if verbatim is None:
                return None, None

            if not isinstance(verbatim, str) and not isinstance(verbatim, int):
                reason = date_name + " date was not a string and is almost certainly wrong: " + str(verbatim)
                report_column_problem(f"problematic {date_name} date", [date_column_name], reason)
                return None, None
            elif isinstance(verbatim, int) or len(verbatim) != 0:
                parsed = parseDate(verbatim)
                parsed, problems = validateResult(parsed, verbatim)
                if len(problems) != 0:
                    reason = problems
                    report_column_problem(f"validation failed for {date_name} date", [date_column_name], reason)
                    return None, None
                return verbatim, parsed
            else:
                return verbatim, None

        captureDateParsed = handleDate('capture', "Date of Capture")
        adjudicationDateParsed = handleDate('adjudication', "Date of Adjudication")

        image_names = []
        for column in ['ImageName0', 'ImageName1', 'ImageName2']:
            if column in schema:
                name = row_values[schema[column]]
                if name is not None:
                    if not IMAGE_NAME_PATTERN.match(name):
                        print(f"[Discarding column] [row {str(row_number)}] Image name {name} does not conform to the image file name pattern.")
                        warnings_kept.add(row_number)
                        warn_reasons["not an image name"] += 1
                        continue
                    image_names.append(name)

        if len(image_names) != 0:
            # TODO: Add to image concordance
            #print(image_names)
            pass

        transcriber_note = None
        if 'Transcriber Note' in schema:
            transcriber_note = row_values[schema['Transcriber Note']]

        importBatch.insert(volume, start_page, end_page, case_name, red_numbers, court,
            row_values[schema["Mixed Commission?"]], ocean,
            adjudicationDateParsed,
            captureDateParsed,
            row_values[schema["Capturing ship/captor"]],
            row_values[schema["Place of Capture"]],
            total, men, women, boys, girls,
            row_values[schema["Register?"]],
            row_values[schema["Notes"]],
            transcriber_note)
    
        num_imported += 1

warnings_kept = [rown for rown in warnings_kept if rown not in discarded]

print("## Summary")
print("- " + str(len(discarded)) + " rows had critical errors and were discarded.")
print("- " + str(num_empty) + " rows were empty.")
print("- " + "Successfully imported " + str(num_imported) + " rows.")
print("- " + str(len(warnings_kept)) + " of those rows generated warnings but were kept.")
#print(courts)
#print(commissions)
print("")

if warnings_kept:
    warn_rows = list(warnings_kept)
    warn_rows.sort();
    print("### Warning Reason Summary")
    #print("The following rows generated warnings while importing but were kept in the data set:")
    #print(warn_rows)
    for (warning, count) in warn_reasons.most_common():
        print(f"- (reported {count} times) {warning}")

if discarded:
    error_rows = list(discarded)
    error_rows.sort();
    print("### Critical Error Summary")
    #print("The following rows failed to import properly and were DISCARDED from the data set:")
    #print(error_rows)
    print("Discard reason summary:")
    for (discard_reason, count) in discard_reasons.most_common():
        print(f"- {discard_reason}, reported {count} times")

# TODO
#print("Values for the \"total\" column that could not be parsed:")
#print(totalNonIntegers)

import os

importBatch.connect_and_execute(user=os.getenv('DB_USER', 'root'),
                                        password=os.getenv('DB_PASSWORD', 'SeniorProject2021!'),
                                        host=os.getenv('DB_HOST', 'localhost'),
                                        port=int(os.getenv('DB_PORT', '3306')),
                                        database=os.getenv('DB_NAME', 'SeniorProject2'))
